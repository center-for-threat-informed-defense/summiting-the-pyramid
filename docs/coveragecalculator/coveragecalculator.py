#!/usr/bin/env python
"""Build score and ATT&CK implementation tables for a Sigma analytic YAML."""

from __future__ import annotations

import argparse
import csv
import fnmatch
import glob
import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any


BUNDLED_SITE_PACKAGES = (
    Path.home()
    / ".cache"
    / "codex-runtimes"
    / "codex-primary-runtime"
    / "dependencies"
    / "python"
    / "Lib"
    / "site-packages"
)

if BUNDLED_SITE_PACKAGES.exists():
    sys.path.insert(0, str(BUNDLED_SITE_PACKAGES))

try:
    import yaml
except ImportError as exc:
    raise SystemExit("Missing dependency: PyYAML. Install it with `pip install pyyaml`.") from exc

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl. Install it with `pip install openpyxl`.") from exc


SCORING_REQUIRED_COLUMNS = {
    "Normalized Field Name",
    "Robustness Score",
    "Precision Score",
}

SCORING_CATEGORY_COLUMNS = (
    "Sigma Category",
    "Sigma logsource_category",
    "Sigma logsource category",
    "logsource_category",
    "logsource category",
)
SCORING_SERVICE_COLUMNS = (
    "Sigma Service",
    "Sigma logsource_service",
    "Sigma logsource_service_",
    "Sigma logsource service",
    "logsource_service",
    "logsource_service_",
    "logsource service",
)
SCORING_PRODUCT_COLUMNS = (
    "Sigma Product",
    "Sigma logsource_product",
    "Sigma logsource product",
    "logsource_product",
    "logsource product",
)
SCORING_EVENT_ID_COLUMNS = (
    "EventID",
    "Event ID",
    "Event Id",
)

ATTACK_REQUIRED_COLUMNS = {
    "ATT&CK Technique ID",
    "ATT&CK Technique Name",
    "Technique Implementation",
    "Implementation Step",
    "System Interaction",
    "ATT&CK Data Component (v19)",
    "Detection Observable",
}

EVENT_FIELD_REQUIRED_COLUMNS = {
    "Log Source",
    "EventID",
    "ATT&CK Data Component (v19)",
    "Message Summary",
    "Field",
    "OCSF Schema Mapping",
}

IMPLEMENTATION_HEADERS = [
    "ATT&CK Technique ID",
    "ATT&CK Technique Name",
    "Technique Implementation name",
    "Implementation Step Number",
    "Implementation Step",
    "System Interaction",
    "ATT&CK Data Component (v19)",
    "Detection Observables",
    "Covered By Analytics",
    "Matched Implementation Count",
    "Total Implementation Count",
    "Implementation Coverage Ratio",
    "Implementation Coverage Percent",
    "Analytic Files",
    "Analytic Titles",
]

FIELD_SCORE_HEADERS = [
    "Analytic File",
    "Analytic Title",
    "Analytic Field",
    "Field Role",
    "Rule Logsource Category",
    "Rule Logsource Service",
    "Rule Logsource Product",
    "Rule EventIDs",
    "Matched Sigma Logsource Category",
    "Matched Sigma Service",
    "Matched Sigma Product",
    "Matched Field",
    "Matched Log Source",
    "Matched EventID",
    "Data Component",
    "OCSF Mapping",
    "Robustness Score",
    "Precision Score",
    "Matched",
]

OCSF_CONTEXT_HEADERS = [
    "Analytic File",
    "Analytic Title",
    "Log Source",
    "EventID",
    "ATT&CK Data Component (v19)",
    "Message Summary",
    "Field",
    "OCSF Schema Mapping",
    "Match Reason",
]

TECHNIQUE_COVERAGE_HEADERS = [
    "ATT&CK Technique ID",
    "ATT&CK Technique Name",
    "Matched Implementation Count",
    "Total Implementation Count",
    "Implementation Coverage Ratio",
    "Implementation Coverage Percent",
    "Matched Technique Implementations",
    "All Technique Implementations",
    "Analytic Files",
    "Analytic Titles",
    "Tagged Analytic Files",
    "Tagged Analytic Titles",
]

ANALYTIC_SCORE_HEADERS = [
    "Analytic File",
    "Analytic Title",
    "ATT&CK Technique ID",
    "ATT&CK Technique Name",
    "Matched Implementation Count",
    "Total Implementation Count",
    "Implementation Coverage Ratio",
    "Implementation Coverage Percent",
    "Matched Technique Implementations",
    "All Technique Implementations",
    "Detection Condition",
    "Robustness Score",
    "Precision Score",
    "Evaluation Expression",
    "Missing Scores",
    "Notes",
]


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(
        description=(
            "Create implementation and field-score tables from a Sigma analytic YAML. "
            "By default, supporting workbooks are loaded from the same directory as this script."
        )
    )
    parser.add_argument(
        "analytic_yaml_args",
        nargs="*",
        help=(
            "One or more Sigma YAML files, directories, or glob patterns. "
            "If only filenames are supplied, the current directory is checked first."
        ),
    )
    parser.add_argument(
        "--analytic-yaml",
        dest="analytic_yaml_opts",
        action="append",
        help="Path to a Sigma YAML file. Can be used multiple times.",
    )
    parser.add_argument(
        "--scoring-workbook",
        default=str(script_dir / "scoring_dictionary.xlsx"),
        help="Path to scoring_dictionary.xlsx. Defaults to the copy next to this script.",
    )
    parser.add_argument(
        "--attack-workbook",
        "--implementation-catalog",
        dest="attack_workbook",
        default=str(script_dir / "implementation_catalog_with_attack_components.xlsx"),
        help="Path to implementation_catalog_with_attack_components.xlsx.",
    )
    parser.add_argument(
        "--event-field-data",
        "--ocsf-mapping-workbook",
        dest="event_field_data",
        default=str(script_dir / "mappings.xlsx"),
        help="Path to the Windows OCSF mappings workbook or CSV. Defaults to mappings.xlsx next to this script.",
    )
    parser.add_argument(
        "--out-prefix",
        default=None,
        help=(
            "Output path prefix. Defaults to outputs/consolidated_analysis next to this script."
        ),
    )
    parser.add_argument(
        "--scoring-sheet",
        default="New Scores",
        help="Scoring worksheet name. Defaults to 'New Scores'.",
    )
    parser.add_argument(
        "--attack-sheet",
        default=None,
        help="ATT&CK worksheet. Defaults to the first sheet with required implementation columns.",
    )
    parser.add_argument(
        "--event-field-sheet",
        default="OCSF Mappings",
        help="Windows OCSF mapping worksheet. Defaults to 'OCSF Mappings'.",
    )
    parser.add_argument(
        "--github-token",
        default=os.environ.get("GITHUB_TOKEN"),
        help="Optional GitHub token. Defaults to the GITHUB_TOKEN environment variable.",
    )
    args = parser.parse_args()

    analytic_inputs = list(args.analytic_yaml_opts or []) + list(args.analytic_yaml_args)
    if not analytic_inputs:
        parser.error("provide at least one Sigma YAML path, for example: python sigma_analytic_tables.py rule.yml")

    github_download_dir = script_dir / ".github_downloads"
    analytic_paths = expand_analytic_inputs(
        analytic_inputs,
        base_dir=script_dir,
        github_download_dir=github_download_dir,
        github_token=args.github_token,
    )
    if not analytic_paths:
        parser.error("no Sigma YAML files were found from the supplied input")

    for analytic_path in analytic_paths:
        require_existing_file(analytic_path, "analytic YAML")

    scoring_workbook = resolve_input_path(args.scoring_workbook, base_dir=script_dir)
    attack_workbook = resolve_input_path(args.attack_workbook, base_dir=script_dir)
    event_field_data = resolve_input_path(args.event_field_data, base_dir=script_dir) if args.event_field_data else None

    require_existing_file(scoring_workbook, "scoring workbook")
    require_existing_file(attack_workbook, "implementation catalog")
    if event_field_data:
        require_existing_file(event_field_data, "OCSF mapping workbook")

    out_prefix = (
        Path(args.out_prefix)
        if args.out_prefix
        else script_dir / "outputs" / "consolidated_analysis"
    )
    out_prefix.parent.mkdir(parents=True, exist_ok=True)

    scoring_rows, scoring_sheet = load_xlsx_rows(
        scoring_workbook, args.scoring_sheet, SCORING_REQUIRED_COLUMNS
    )
    attack_rows, attack_sheet = load_xlsx_rows(attack_workbook, args.attack_sheet, ATTACK_REQUIRED_COLUMNS)

    event_rows: list[dict[str, Any]] = []
    event_source = ""
    if event_field_data:
        event_rows, event_source = load_event_field_rows(event_field_data, args.event_field_sheet)

    analytic_summaries: list[dict[str, Any]] = []
    all_event_context_rows: list[dict[str, Any]] = []
    all_field_score_rows: list[dict[str, Any]] = []
    all_implementation_rows: list[dict[str, Any]] = []
    all_analytic_score_rows: list[dict[str, Any]] = []

    for analytic_path in analytic_paths:
        rule = extract_sigma_rule(analytic_path)
        event_context_rows = build_event_context_rows(rule, event_rows)
        field_score_rows = build_field_score_rows(rule, scoring_rows)
        analytic_score_row = build_analytic_score_row(rule, field_score_rows)
        data_component_candidates = build_data_component_candidates(rule, field_score_rows, event_context_rows)
        implementation_rows = build_implementation_rows(rule, attack_rows, data_component_candidates)
        analytic_score_rows = build_analytic_score_rows(rule, analytic_score_row, attack_rows, implementation_rows)

        all_event_context_rows.extend(event_context_rows)
        all_field_score_rows.extend(field_score_rows)
        all_implementation_rows.extend(implementation_rows)
        all_analytic_score_rows.extend(analytic_score_rows)
        analytic_summaries.append(
            {
                "rule": rule,
                "analytic_score": analytic_score_row,
                "attack_data_component_candidates": data_component_candidates,
                "event_context_match_count": len(event_context_rows),
                "implementation_match_count": len(implementation_rows),
                "field_score_row_count": len(field_score_rows),
                "field_score_match_count": sum(1 for row in field_score_rows if row["Matched"] == "yes"),
            }
        )

    matched_implementation_rows = consolidate_implementation_rows(all_implementation_rows)
    technique_coverage_rows = build_technique_coverage_rows(analytic_summaries, attack_rows, matched_implementation_rows)
    implementation_rows = build_implementation_catalog_rows(
        analytic_summaries,
        attack_rows,
        matched_implementation_rows,
        technique_coverage_rows,
    )
    summary_rows = build_summary_rows(
        analytic_summaries,
        scoring_sheet,
        attack_sheet,
        event_source,
        all_event_context_rows,
        all_field_score_rows,
        implementation_rows,
        technique_coverage_rows,
        all_analytic_score_rows,
    )

    workbook_path = out_prefix.with_suffix(".xlsx")

    write_workbook(
        workbook_path,
        implementation_rows,
        all_field_score_rows,
        all_event_context_rows,
        technique_coverage_rows,
        all_analytic_score_rows,
        summary_rows,
    )
    print(f"Analytics processed: {len(analytic_summaries)}")
    print(f"Implementation catalog rows: {len(implementation_rows)}")
    print(f"Covered implementation rows: {sum(1 for row in implementation_rows if row.get('Covered By Analytics') == 'yes')}")
    print(f"Technique coverage rows: {len(technique_coverage_rows)}")
    print(f"Field score rows: {len(all_field_score_rows)}")
    print(f"Field score matches: {sum(1 for row in all_field_score_rows if row['Matched'] == 'yes')}")
    print(f"Event context matches: {len(all_event_context_rows)}")
    for summary in analytic_summaries:
        rule = summary["rule"]
        analytic_score = format_summary_score_value(summary["analytic_score"])
        print(
            f"- {rule['analytic_file']}: {summary['field_score_match_count']}/"
            f"{summary['field_score_row_count']} field scores, "
            f"{summary['implementation_match_count']} implementation matches, "
            f"score {analytic_score}"
        )
    print(f"Wrote workbook: {workbook_path}")


def expand_analytic_inputs(
    input_texts: list[str],
    base_dir: Path,
    github_download_dir: Path,
    github_token: str | None,
) -> list[Path]:
    paths: list[Path] = []
    for input_text in input_texts:
        expanded_paths = expand_one_analytic_input(input_text, base_dir, github_download_dir, github_token)
        if not expanded_paths:
            paths.append(resolve_input_path(input_text, base_dir=base_dir))
        else:
            paths.extend(expanded_paths)

    seen: set[str] = set()
    unique_paths: list[Path] = []
    for path in paths:
        key = str(path.resolve()).lower() if path.exists() else str(path).lower()
        if key in seen:
            continue
        seen.add(key)
        unique_paths.append(path)
    return unique_paths


def expand_one_analytic_input(
    input_text: str,
    base_dir: Path,
    github_download_dir: Path,
    github_token: str | None,
) -> list[Path]:
    if is_url(input_text):
        if is_github_url(input_text):
            github_paths = expand_github_input(input_text, github_download_dir, github_token)
            if not github_paths:
                raise SystemExit(f"No YAML files found at GitHub URL: {input_text}")
            return github_paths
        raise SystemExit(f"Unsupported URL input. Only GitHub URLs are supported: {input_text}")

    if has_glob_chars(input_text):
        matches: list[Path] = []
        candidate_patterns = [input_text]
        raw_path = Path(input_text)
        if not raw_path.is_absolute():
            candidate_patterns.append(str(base_dir / raw_path))

        for pattern in candidate_patterns:
            matches.extend(Path(match).resolve() for match in glob.glob(pattern))
        return sorted(path for path in matches if is_yaml_path(path))

    path = resolve_input_path(input_text, base_dir=base_dir)
    if path.is_dir():
        return sorted(
            [candidate for candidate in path.iterdir() if candidate.is_file() and is_yaml_path(candidate)]
        )
    return [path] if is_yaml_path(path) else []


def is_url(path_text: str) -> bool:
    parsed = urllib.parse.urlparse(path_text)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def is_github_url(path_text: str) -> bool:
    host = urllib.parse.urlparse(path_text).netloc.lower()
    return host in {"github.com", "www.github.com", "raw.githubusercontent.com"}


def has_glob_chars(path_text: str) -> bool:
    return any(char in path_text for char in "*?[]")


def is_yaml_path(path: Path) -> bool:
    return path.suffix.lower() in {".yml", ".yaml"}


def expand_github_input(url: str, download_dir: Path, token: str | None) -> list[Path]:
    parsed = urllib.parse.urlparse(url)
    host = parsed.netloc.lower()
    if host == "raw.githubusercontent.com":
        return download_github_raw_url(url, download_dir, token)

    owner, repo, kind, ref, repo_path = parse_github_url(url, token)
    if kind == "blob":
        item = github_api_json(github_contents_url(owner, repo, repo_path, ref), token)
        if not isinstance(item, dict) or item.get("type") != "file":
            raise SystemExit(f"GitHub blob URL did not resolve to a file: {url}")
        return download_github_file_item(item, owner, repo, ref, download_dir, token)

    if kind in {"tree", "repo"}:
        return download_github_directory(owner, repo, ref, repo_path, download_dir, token)

    raise SystemExit(f"Unsupported GitHub URL: {url}")


def parse_github_url(url: str, token: str | None) -> tuple[str, str, str, str, str]:
    parsed = urllib.parse.urlparse(url)
    parts = [urllib.parse.unquote(part) for part in parsed.path.strip("/").split("/") if part]
    if len(parts) < 2:
        raise SystemExit(f"GitHub URL must include owner and repo: {url}")

    owner, repo = parts[0], strip_git_suffix(parts[1])
    if len(parts) == 2:
        default_branch = get_github_default_branch(owner, repo, token)
        return owner, repo, "repo", default_branch, ""

    marker = parts[2]
    if marker not in {"blob", "tree"}:
        default_branch = get_github_default_branch(owner, repo, token)
        repo_path = "/".join(parts[2:])
        return owner, repo, "repo", default_branch, repo_path

    remainder = parts[3:]
    if not remainder:
        default_branch = get_github_default_branch(owner, repo, token)
        return owner, repo, marker, default_branch, ""

    ref, repo_path = resolve_github_ref_and_path(owner, repo, marker, remainder, token)
    return owner, repo, marker, ref, repo_path


def resolve_github_ref_and_path(
    owner: str,
    repo: str,
    kind: str,
    remainder: list[str],
    token: str | None,
) -> tuple[str, str]:
    expected_type = "file" if kind == "blob" else "dir"

    # Try longer refs first so branch names containing slashes can still resolve.
    for split_index in range(len(remainder), 0, -1):
        candidate_ref = "/".join(remainder[:split_index])
        candidate_path = "/".join(remainder[split_index:])
        try:
            item = github_api_json(github_contents_url(owner, repo, candidate_path, candidate_ref), token)
        except SystemExit:
            continue

        if kind == "tree" and candidate_path == "" and isinstance(item, list):
            return candidate_ref, candidate_path
        if isinstance(item, dict) and item.get("type") == expected_type:
            return candidate_ref, candidate_path
        if kind == "tree" and isinstance(item, list):
            return candidate_ref, candidate_path

    # Fall back to the common case: first path segment is the branch/tag.
    return remainder[0], "/".join(remainder[1:])


def get_github_default_branch(owner: str, repo: str, token: str | None) -> str:
    repo_info = github_api_json(f"https://api.github.com/repos/{owner}/{repo}", token)
    if isinstance(repo_info, dict) and repo_info.get("default_branch"):
        return clean_cell(repo_info["default_branch"])
    raise SystemExit(f"Could not determine default branch for {owner}/{repo}.")


def download_github_directory(
    owner: str,
    repo: str,
    ref: str,
    repo_path: str,
    download_dir: Path,
    token: str | None,
) -> list[Path]:
    item = github_api_json(github_contents_url(owner, repo, repo_path, ref), token)
    if isinstance(item, dict) and item.get("type") == "file":
        return download_github_file_item(item, owner, repo, ref, download_dir, token)

    if not isinstance(item, list):
        raise SystemExit(f"GitHub URL did not resolve to a file or directory: {owner}/{repo}/{repo_path}")

    downloaded: list[Path] = []
    for child in item:
        if not isinstance(child, dict):
            continue
        if child.get("type") == "dir":
            downloaded.extend(download_github_directory(owner, repo, ref, clean_cell(child.get("path")), download_dir, token))
        elif child.get("type") == "file" and is_yaml_name(clean_cell(child.get("name"))):
            downloaded.extend(download_github_file_item(child, owner, repo, ref, download_dir, token))
    return sorted(downloaded)


def download_github_file_item(
    item: dict[str, Any],
    owner: str,
    repo: str,
    ref: str,
    download_dir: Path,
    token: str | None,
) -> list[Path]:
    name = clean_cell(item.get("name"))
    if not is_yaml_name(name):
        return []

    download_url = clean_cell(item.get("download_url"))
    if not download_url:
        repo_path = clean_cell(item.get("path"))
        download_url = raw_github_url(owner, repo, ref, repo_path)

    repo_path = clean_cell(item.get("path")) or name
    local_path = local_github_download_path(download_dir, owner, repo, ref, repo_path)
    download_url_to_file(download_url, local_path, token)
    return [local_path]


def download_github_raw_url(url: str, download_dir: Path, token: str | None) -> list[Path]:
    parsed = urllib.parse.urlparse(url)
    parts = [urllib.parse.unquote(part) for part in parsed.path.strip("/").split("/") if part]
    if len(parts) < 4:
        raise SystemExit(f"Raw GitHub URL must include owner, repo, ref, and path: {url}")

    owner, repo, ref = parts[0], strip_git_suffix(parts[1]), parts[2]
    repo_path = "/".join(parts[3:])
    if not is_yaml_name(repo_path):
        raise SystemExit(f"Raw GitHub URL is not a YAML file: {url}")

    local_path = local_github_download_path(download_dir, owner, repo, ref, repo_path)
    download_url_to_file(url, local_path, token)
    return [local_path]


def github_contents_url(owner: str, repo: str, repo_path: str, ref: str) -> str:
    encoded_path = "/".join(urllib.parse.quote(part) for part in repo_path.split("/") if part)
    if encoded_path:
        return f"https://api.github.com/repos/{owner}/{repo}/contents/{encoded_path}?ref={urllib.parse.quote(ref, safe='')}"
    return f"https://api.github.com/repos/{owner}/{repo}/contents?ref={urllib.parse.quote(ref, safe='')}"


def github_api_json(url: str, token: str | None) -> Any:
    data = http_get(url, token, accept="application/vnd.github+json")
    return json.loads(data.decode("utf-8"))


def http_get(url: str, token: str | None, accept: str | None = None) -> bytes:
    headers = {"User-Agent": "sigma-analytic-tables"}
    if accept:
        headers["Accept"] = accept
    if token:
        headers["Authorization"] = f"Bearer {token}"

    request = urllib.request.Request(url, headers=headers)
    try:
        with urllib.request.urlopen(request, timeout=60) as response:
            return response.read()
    except urllib.error.HTTPError as exc:
        message = exc.read().decode("utf-8", errors="replace")
        raise SystemExit(f"GitHub request failed ({exc.code}) for {url}: {message}") from exc
    except urllib.error.URLError as exc:
        raise SystemExit(f"GitHub request failed for {url}: {exc.reason}") from exc


def download_url_to_file(url: str, local_path: Path, token: str | None) -> None:
    local_path.parent.mkdir(parents=True, exist_ok=True)
    data = http_get(url, token)
    local_path.write_bytes(data)


def local_github_download_path(download_dir: Path, owner: str, repo: str, ref: str, repo_path: str) -> Path:
    safe_ref = sanitize_path_segment(ref)
    return download_dir / owner / repo / safe_ref / Path(repo_path)


def raw_github_url(owner: str, repo: str, ref: str, repo_path: str) -> str:
    encoded_path = "/".join(urllib.parse.quote(part) for part in repo_path.split("/") if part)
    return f"https://raw.githubusercontent.com/{owner}/{repo}/{urllib.parse.quote(ref, safe='')}/{encoded_path}"


def strip_git_suffix(repo: str) -> str:
    return repo[:-4] if repo.endswith(".git") else repo


def sanitize_path_segment(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._") or "ref"


def is_yaml_name(name: str) -> bool:
    return name.lower().endswith((".yml", ".yaml"))


def resolve_input_path(path_text: str, base_dir: Path | None = None) -> Path:
    path = Path(path_text).expanduser()
    if path.is_absolute():
        return path
    if path.exists():
        return path.resolve()
    if base_dir:
        base_path = base_dir / path
        if base_path.exists():
            return base_path.resolve()
        return base_path
    return path


def require_existing_file(path: Path, label: str) -> None:
    if not path.exists():
        raise SystemExit(f"Missing {label}: {path}")
    if not path.is_file():
        raise SystemExit(f"Expected {label} to be a file: {path}")


def extract_sigma_rule(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        raw = yaml.safe_load(handle)

    if not isinstance(raw, dict):
        raise ValueError(f"{path} did not parse to a YAML object.")

    logsource = raw.get("logsource") or {}
    detection = raw.get("detection") or {}
    if not isinstance(logsource, dict):
        raise ValueError("Expected logsource to be a mapping.")
    if not isinstance(detection, dict):
        raise ValueError("Expected detection to be a mapping.")

    rule_logsource_category = clean_cell(logsource.get("category"))
    rule_logsource_service = clean_cell(logsource.get("service"))
    rule_logsource_product = clean_cell(logsource.get("product"))

    detection_fields: list[str] = []
    detection_filters: list[str] = []
    detection_values: dict[str, list[str]] = {}
    detection_blocks: dict[str, dict[str, Any]] = {}
    detection_block_order: list[str] = []

    for key, value in detection.items():
        if key == "condition":
            continue
        role = classify_detection_key(key)
        detection_block_order.append(key)
        detection_blocks[key] = {
            "role": role,
            "expression": build_detection_block_expression(key, value),
        }
        entries = extract_detection_entries(key, value)
        for field, field_value in entries:
            if field_value != "":
                detection_values.setdefault(field, [])
                append_unique(detection_values[field], field_value)
            if is_event_id_field(field):
                continue
            if role == "filter":
                append_unique(detection_filters, field)
            else:
                append_unique(detection_fields, field)

    tags = raw.get("tags") or []
    if isinstance(tags, str):
        tags = [tags]

    return {
        "analytic_file": path.name,
        "analytic_path": str(path),
        "rule_title": raw.get("title"),
        "rule_logsource_product": rule_logsource_product,
        "rule_logsource_category": rule_logsource_category,
        "rule_logsource_service": rule_logsource_service,
        "rule_tags": [str(tag) for tag in tags],
        "attack_technique_ids": extract_attack_technique_ids(tags),
        "detection_condition": clean_cell(detection.get("condition")),
        "detection_fields": detection_fields,
        "detection_filters": detection_filters,
        "detection_values": detection_values,
        "detection_blocks": detection_blocks,
        "detection_block_order": detection_block_order,
    }


def classify_detection_key(key: str) -> str:
    normalized = key.lower()
    return "filter" if normalized == "filter" or normalized.startswith("filter_") else "field"


def build_detection_block_expression(key: str, value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return combine_expression_nodes(
            "and",
            [field_expression(field_name) for field_name in value.keys()],
        )
    if isinstance(value, list):
        children: list[dict[str, Any]] = []
        for item in value:
            if isinstance(item, dict):
                children.append(
                    combine_expression_nodes(
                        "and",
                        [field_expression(field_name) for field_name in item.keys()],
                    )
                )
            else:
                children.append(field_expression(item))
        return combine_expression_nodes("or", children)

    inferred_name = re.sub(r"^(selection_|filter_)", "", key)
    return field_expression(inferred_name)


def field_expression(raw_name: Any) -> dict[str, Any]:
    field_name = str(raw_name).split("|", 1)[0].strip()
    if is_event_id_field(field_name):
        return {"type": "empty"}
    return {"type": "field", "field": field_name}


def combine_expression_nodes(operator: str, children: list[dict[str, Any]]) -> dict[str, Any]:
    children = [child for child in children if not expression_is_empty(child)]
    if not children:
        return {"type": "empty"}
    if len(children) == 1:
        return children[0]
    return {"type": operator, "children": children}


def expression_is_empty(expression: dict[str, Any]) -> bool:
    expression_type = expression.get("type")
    if expression_type == "empty":
        return True
    if expression_type == "field":
        return not clean_cell(expression.get("field"))
    if expression_type in {"ref", "missing_ref"}:
        return not clean_cell(expression.get("name"))
    if expression_type in {"and", "or"}:
        return not expression.get("children")
    return False


def extract_detection_entries(key: str, value: Any) -> list[tuple[str, str]]:
    entries: list[tuple[str, str]] = []
    if isinstance(value, dict):
        for field_name, field_value in value.items():
            append_detection_entry(entries, field_name, field_value)
    elif isinstance(value, list):
        for item in value:
            if isinstance(item, dict):
                for field_name, field_value in item.items():
                    append_detection_entry(entries, field_name, field_value)
            else:
                append_detection_entry(entries, item, "")
    else:
        inferred_name = re.sub(r"^(selection_|filter_)", "", key)
        append_detection_entry(entries, inferred_name, value)
    return entries


def append_detection_entry(entries: list[tuple[str, str]], raw_name: Any, raw_value: Any) -> None:
    field_name = str(raw_name).split("|", 1)[0].strip()
    if not field_name:
        return
    if isinstance(raw_value, list):
        value_text = "; ".join(clean_cell(value) for value in raw_value)
    elif isinstance(raw_value, dict):
        value_text = json.dumps(raw_value, sort_keys=True)
    else:
        value_text = clean_cell(raw_value)
    entries.append((field_name, value_text))


def append_unique(target: list[str], value: str) -> None:
    if value and value not in target:
        target.append(value)


def extract_attack_technique_ids(tags: list[Any]) -> list[str]:
    technique_ids: list[str] = []
    for tag in tags:
        for match in re.finditer(r"attack\.t(?P<main>\d{4})(?:\.(?P<sub>\d{3}))?", str(tag), re.I):
            technique_id = f"T{match.group('main')}"
            if match.group("sub"):
                technique_id = f"{technique_id}.{match.group('sub')}"
            append_unique(technique_ids, technique_id)
    return technique_ids


def load_xlsx_rows(
    workbook_path: Path, requested_sheet: str | None, required_columns: set[str]
) -> tuple[list[dict[str, Any]], str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = select_sheet(workbook, requested_sheet, required_columns)
    return sheet_to_dicts(worksheet), worksheet.title


def select_sheet(workbook: Any, requested_sheet: str | None, required_columns: set[str]) -> Any:
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise ValueError(f"Workbook does not contain worksheet {requested_sheet!r}.")
        worksheet = workbook[requested_sheet]
        ensure_columns(worksheet, required_columns)
        return worksheet

    for worksheet in workbook.worksheets:
        if required_columns.issubset(set(get_headers(worksheet))):
            return worksheet
    raise ValueError(f"No worksheet contains required columns: {sorted(required_columns)}")


def ensure_columns(worksheet: Any, required_columns: set[str]) -> None:
    headers = set(get_headers(worksheet))
    missing = sorted(required_columns - headers)
    if missing:
        raise ValueError(f"Worksheet {worksheet.title!r} is missing columns: {missing}")


def get_headers(worksheet: Any) -> list[str]:
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [clean_cell(cell) for cell in first_row]


def sheet_to_dicts(worksheet: Any) -> list[dict[str, Any]]:
    headers = get_headers(worksheet)
    rows: list[dict[str, Any]] = []
    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in raw_row):
            continue
        rows.append({headers[index]: raw_row[index] if index < len(raw_row) else None for index in range(len(headers))})
    return rows


def load_event_field_rows(path: Path, requested_sheet: str) -> tuple[list[dict[str, Any]], str]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        rows, sheet = load_xlsx_rows(path, requested_sheet, EVENT_FIELD_REQUIRED_COLUMNS)
        return rows, f"{path.name}:{sheet}"

    rows = load_csv_rows(path)
    missing = sorted(EVENT_FIELD_REQUIRED_COLUMNS - set(rows[0].keys())) if rows else sorted(EVENT_FIELD_REQUIRED_COLUMNS)
    if missing:
        raise ValueError(f"{path} is missing columns: {missing}")
    return rows, path.name


def load_csv_rows(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            with path.open("r", newline="", encoding=encoding) as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("csv", b"", 0, 1, f"Could not decode {path}")


def build_event_context_rows(rule: dict[str, Any], event_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not event_rows:
        return []

    event_ids = set(rule_event_ids(rule))
    detection_fields = set(rule["detection_fields"]) | set(rule["detection_filters"])

    matches: list[dict[str, Any]] = []
    for row in event_rows:
        if not logsource_matches(rule, row.get("Log Source")):
            continue

        event_id_matches = clean_cell(row.get("EventID")) in event_ids if event_ids else False
        field_matches = field_matches_any(row.get("Field"), detection_fields)
        reasons: list[str] = []
        if event_ids:
            if not (event_id_matches and field_matches):
                continue
            reasons.extend(["EventID", "Field"])
        elif field_matches:
            reasons.append("Field")
        if not reasons:
            continue

        matches.append(
            {
                "Analytic File": rule["analytic_file"],
                "Analytic Title": rule["rule_title"],
                "Log Source": clean_cell(row.get("Log Source")),
                "EventID": clean_cell(row.get("EventID")),
                "ATT&CK Data Component (v19)": clean_cell(row.get("ATT&CK Data Component (v19)")),
                "Message Summary": clean_cell(row.get("Message Summary")),
                "Field": clean_cell(row.get("Field")),
                "OCSF Schema Mapping": clean_cell(row.get("OCSF Schema Mapping")),
                "Match Reason": ", ".join(reasons),
            }
        )
    return matches


def build_field_score_rows(rule: dict[str, Any], scoring_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output_rows: list[dict[str, Any]] = []
    seen_output_rows: set[tuple[str, ...]] = set()
    field_items = [("detection_field", field) for field in rule["detection_fields"]]
    field_items.extend(("detection_filter", field) for field in rule["detection_filters"])

    for role, field in field_items:
        context_matches = [row for row in scoring_rows if scoring_context_matches(rule, row)]
        matches = [row for row in context_matches if same_text(field, row.get("Normalized Field Name"))]
        if matches:
            for match in matches:
                append_unique_score_row(
                    output_rows,
                    seen_output_rows,
                    field_score_output_row(rule, role, field, match, matched=True),
                )
        else:
            append_unique_score_row(
                output_rows,
                seen_output_rows,
                field_score_output_row(rule, role, field, {}, matched=False),
            )
    return output_rows


def append_unique_score_row(
    output_rows: list[dict[str, Any]],
    seen_output_rows: set[tuple[str, ...]],
    row: dict[str, Any],
) -> None:
    row_key = tuple(clean_cell(row.get(header)) for header in FIELD_SCORE_HEADERS)
    if row_key in seen_output_rows:
        return
    seen_output_rows.add(row_key)
    output_rows.append(row)


def field_score_output_row(rule: dict[str, Any], role: str, field: str, match: dict[str, Any], matched: bool) -> dict[str, Any]:
    return {
        "Analytic File": rule["analytic_file"],
        "Analytic Title": rule["rule_title"],
        "Analytic Field": field,
        "Field Role": role,
        "Rule Logsource Category": rule["rule_logsource_category"],
        "Rule Logsource Service": rule["rule_logsource_service"],
        "Rule Logsource Product": rule["rule_logsource_product"],
        "Rule EventIDs": "; ".join(rule_event_ids(rule)),
        "Matched Sigma Logsource Category": scoring_row_value(match, SCORING_CATEGORY_COLUMNS),
        "Matched Sigma Service": scoring_row_value(match, SCORING_SERVICE_COLUMNS),
        "Matched Sigma Product": scoring_row_value(match, SCORING_PRODUCT_COLUMNS),
        "Matched Field": clean_cell(match.get("Normalized Field Name")),
        "Matched Log Source": clean_cell(match.get("Log Source")),
        "Matched EventID": clean_cell(match.get("EventID")),
        "Data Component": clean_cell(match.get("Data Component")),
        "OCSF Mapping": clean_cell(match.get("OCSF Mapping") or match.get("OCSF Schema Mapping")),
        "Robustness Score": clean_cell(match.get("Robustness Score")),
        "Precision Score": clean_cell(match.get("Precision Score")),
        "Matched": "yes" if matched else "no",
    }


def build_analytic_score_row(rule: dict[str, Any], field_score_rows: list[dict[str, Any]]) -> dict[str, Any]:
    notes: list[str] = []
    field_scores, field_score_notes = build_field_score_lookup(field_score_rows)
    notes.extend(field_score_notes)

    try:
        scoreable_detection_blocks = {
            block_name: block
            for block_name, block in rule["detection_blocks"].items()
            if not expression_is_empty(block.get("expression", {"type": "empty"}))
        }
        block_scores = {
            block_name: evaluate_score_expression(block["expression"], field_scores)
            for block_name, block in scoreable_detection_blocks.items()
        }
        condition_expression = parse_detection_condition(rule)
        result = evaluate_score_expression(condition_expression, field_scores, block_scores)
        expression_text = score_expression_to_text(condition_expression, rule["detection_blocks"])
    except ValueError as exc:
        result = {"robustness": None, "precision": None, "missing": [], "notes": [str(exc)]}
        expression_text = ""

    notes.extend(result.get("notes", []))
    missing_scores = sorted(set(result.get("missing", [])))
    if missing_scores:
        notes.append("Analytic tuple not calculated because one or more referenced fields/blocks are unscored.")

    robustness = result.get("robustness")
    precision = result.get("precision")
    return {
        "Analytic File": rule["analytic_file"],
        "Analytic Title": rule["rule_title"],
        "Detection Condition": rule["detection_condition"],
        "Robustness Score": clean_cell(robustness),
        "Precision Score": clean_cell(precision),
        "Evaluation Expression": expression_text,
        "Missing Scores": "; ".join(missing_scores),
        "Notes": "; ".join(dedupe_preserve_order(notes)),
    }


def build_analytic_score_rows(
    rule: dict[str, Any],
    analytic_score_row: dict[str, Any],
    attack_rows: list[dict[str, Any]],
    implementation_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    technique_ids = rule["attack_technique_ids"]
    if not technique_ids:
        return [analytic_score_coverage_row(analytic_score_row, "", "", [], [])]

    output_rows: list[dict[str, Any]] = []
    for technique_id in technique_ids:
        technique_name, all_implementations = technique_catalog_details(technique_id, attack_rows)
        matched_implementations: list[str] = []
        for row in implementation_rows:
            if clean_cell(row.get("ATT&CK Technique ID")) != technique_id:
                continue
            if not technique_name and clean_cell(row.get("ATT&CK Technique Name")):
                technique_name = clean_cell(row.get("ATT&CK Technique Name"))
            append_unique(matched_implementations, clean_cell(row.get("Technique Implementation name")))
        output_rows.append(
            analytic_score_coverage_row(
                analytic_score_row,
                technique_id,
                technique_name,
                matched_implementations,
                all_implementations,
            )
        )
    return output_rows


def technique_catalog_details(technique_id: str, attack_rows: list[dict[str, Any]]) -> tuple[str, list[str]]:
    technique_name = ""
    implementations: list[str] = []
    for row in attack_rows:
        if clean_cell(row.get("ATT&CK Technique ID")) != technique_id:
            continue
        if not technique_name and clean_cell(row.get("ATT&CK Technique Name")):
            technique_name = clean_cell(row.get("ATT&CK Technique Name"))
        append_unique(implementations, clean_cell(row.get("Technique Implementation")))
    return technique_name, implementations


def analytic_score_coverage_row(
    analytic_score_row: dict[str, Any],
    technique_id: str,
    technique_name: str,
    matched_implementations: list[str],
    all_implementations: list[str],
) -> dict[str, Any]:
    matched_count = len(matched_implementations)
    total_count = len(all_implementations)
    row = dict(analytic_score_row)
    row.update(
        {
            "ATT&CK Technique ID": technique_id,
            "ATT&CK Technique Name": technique_name,
            "Matched Implementation Count": clean_cell(matched_count) if technique_id else "",
            "Total Implementation Count": clean_cell(total_count) if technique_id else "",
            "Implementation Coverage Ratio": f"{matched_count}/{total_count}" if technique_id else "",
            "Implementation Coverage Percent": coverage_percent(matched_count, total_count) if technique_id else "",
            "Matched Technique Implementations": "; ".join(matched_implementations),
            "All Technique Implementations": "; ".join(all_implementations),
        }
    )
    return row


def build_field_score_lookup(field_score_rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], list[str]]:
    scores_by_field: dict[str, list[dict[str, int]]] = {}
    notes: list[str] = []
    for row in field_score_rows:
        if row.get("Matched") != "yes":
            continue
        field = normalize_field_name(row.get("Analytic Field"))
        if is_event_id_field(field):
            continue
        robustness = parse_score_value(row.get("Robustness Score"))
        precision = parse_score_value(row.get("Precision Score"))
        if not field or robustness is None or precision is None:
            continue
        scores_by_field.setdefault(field, []).append({"robustness": robustness, "precision": precision})

    field_scores: dict[str, dict[str, Any]] = {}
    for field, scores in scores_by_field.items():
        if len(scores) == 1:
            field_scores[field] = {**scores[0], "missing": [], "notes": []}
        else:
            field_scores[field] = combine_score_results("or", scores)
            notes.append(f"Field {field} had multiple score rows and was consolidated as OR.")
    return field_scores, notes


def parse_score_value(value: Any) -> int | None:
    text = clean_cell(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_detection_condition(rule: dict[str, Any]) -> dict[str, Any]:
    condition = clean_cell(rule.get("detection_condition"))
    all_block_order = rule.get("detection_block_order") or []
    detection_blocks = rule.get("detection_blocks") or {}
    block_order = [
        block_name
        for block_name in all_block_order
        if not expression_is_empty(detection_blocks.get(block_name, {}).get("expression", {"type": "empty"}))
    ]
    ignored_block_names = {block_name for block_name in all_block_order if block_name not in set(block_order)}
    if not condition:
        if len(block_order) == 1:
            return {"type": "ref", "name": block_order[0]}
        if block_order:
            return combine_expression_nodes("and", [{"type": "ref", "name": name} for name in block_order])
        raise ValueError("Detection condition is empty and no detection blocks were found.")
    parser = ConditionParser(condition, block_order, ignored_block_names)
    return parser.parse()


class ConditionParser:
    TOKEN_PATTERN = re.compile(r"\b(?:all|any|1|of|them|and|or|not)\b|[()]|[A-Za-z0-9_*?.-]+", re.I)

    def __init__(self, condition: str, block_names: list[str], ignored_block_names: set[str] | None = None):
        self.condition = condition
        self.tokens = self.TOKEN_PATTERN.findall(condition)
        self.block_names = block_names
        self.ignored_block_names = ignored_block_names or set()
        self.index = 0

    def parse(self) -> dict[str, Any]:
        if not self.tokens:
            raise ValueError("Detection condition did not contain parseable tokens.")
        expression = self.parse_or()
        if self.peek() is not None:
            raise ValueError(f"Unexpected token in detection condition: {self.peek()}")
        return expression

    def parse_or(self) -> dict[str, Any]:
        children = [self.parse_and()]
        while self.accept("or"):
            children.append(self.parse_and())
        return combine_expression_nodes("or", children)

    def parse_and(self) -> dict[str, Any]:
        children = [self.parse_unary()]
        while self.accept("and"):
            children.append(self.parse_unary())
        return combine_expression_nodes("and", children)

    def parse_unary(self) -> dict[str, Any]:
        if self.accept("not"):
            return {"type": "not", "child": self.parse_unary()}
        return self.parse_primary()

    def parse_primary(self) -> dict[str, Any]:
        token = self.peek()
        if token is None:
            raise ValueError("Unexpected end of detection condition.")
        if token == "(":
            self.index += 1
            expression = self.parse_or()
            if not self.accept(")"):
                raise ValueError("Missing closing parenthesis in detection condition.")
            return expression

        lower_token = token.lower()
        if lower_token in {"all", "1", "any"}:
            self.index += 1
            if not self.accept("of"):
                raise ValueError(f"Expected 'of' after {token!r} in detection condition.")
            pattern = self.consume()
            if pattern is None:
                raise ValueError(f"Expected detection selector after {token!r} of.")
            operator = "and" if lower_token == "all" else "or"
            return self.expand_selector(pattern, operator)

        self.index += 1
        if lower_token == "them" or has_glob_chars(token):
            return self.expand_selector(token, "or")
        if token in self.ignored_block_names:
            return {"type": "empty"}
        return {"type": "ref", "name": token}

    def expand_selector(self, pattern: str, operator: str) -> dict[str, Any]:
        lower_pattern = pattern.lower()
        ignored_matches: list[str] = []
        if lower_pattern == "them":
            matches = self.block_names
            ignored_matches = list(self.ignored_block_names)
        elif has_glob_chars(pattern):
            matches = [name for name in self.block_names if fnmatch.fnmatchcase(name, pattern)]
            ignored_matches = [name for name in self.ignored_block_names if fnmatch.fnmatchcase(name, pattern)]
        else:
            matches = [pattern] if pattern in self.block_names else []
            ignored_matches = [pattern] if pattern in self.ignored_block_names else []
        if not matches:
            if ignored_matches:
                return {"type": "empty"}
            return {"type": "missing_ref", "name": pattern}
        return combine_expression_nodes(operator, [{"type": "ref", "name": name} for name in matches])

    def peek(self) -> str | None:
        return self.tokens[self.index] if self.index < len(self.tokens) else None

    def consume(self) -> str | None:
        token = self.peek()
        if token is not None:
            self.index += 1
        return token

    def accept(self, expected: str) -> bool:
        token = self.peek()
        if token is not None and token.lower() == expected.lower():
            self.index += 1
            return True
        return False


def evaluate_score_expression(
    expression: dict[str, Any],
    field_scores: dict[str, dict[str, Any]],
    block_scores: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    expression_type = expression.get("type")
    if expression_type == "field":
        field = normalize_field_name(expression.get("field"))
        if field in field_scores:
            return field_scores[field]
        return missing_score_result(field)
    if expression_type == "ref":
        block_name = clean_cell(expression.get("name"))
        if block_scores and block_name in block_scores:
            return block_scores[block_name]
        return missing_score_result(block_name)
    if expression_type == "missing_ref":
        return missing_score_result(clean_cell(expression.get("name")))
    if expression_type == "not":
        return evaluate_score_expression(expression.get("child", {"type": "empty"}), field_scores, block_scores)
    if expression_type in {"and", "or"}:
        child_results = [
            evaluate_score_expression(child, field_scores, block_scores)
            for child in expression.get("children", [])
        ]
        return combine_score_results(expression_type, child_results)
    return {"robustness": None, "precision": None, "missing": ["<empty expression>"], "notes": []}


def missing_score_result(name: str) -> dict[str, Any]:
    return {"robustness": None, "precision": None, "missing": [name or "<unknown>"], "notes": []}


def combine_score_results(operator: str, child_results: list[dict[str, Any]]) -> dict[str, Any]:
    missing = sorted({item for result in child_results for item in result.get("missing", [])})
    notes = dedupe_preserve_order([note for result in child_results for note in result.get("notes", [])])
    if missing or not child_results:
        return {"robustness": None, "precision": None, "missing": missing or ["<empty expression>"], "notes": notes}

    robustness_values = [int(result["robustness"]) for result in child_results]
    precision_values = [int(result["precision"]) for result in child_results]
    if operator == "and":
        robustness = min(robustness_values)
        precision = max(precision_values)
    elif operator == "or":
        robustness = max(robustness_values)
        precision = min(precision_values)
    else:
        raise ValueError(f"Unsupported score operator: {operator}")
    return {"robustness": robustness, "precision": precision, "missing": [], "notes": notes}


def score_expression_to_text(expression: dict[str, Any], detection_blocks: dict[str, dict[str, Any]] | None = None) -> str:
    expression_type = expression.get("type")
    if expression_type == "field":
        return clean_cell(expression.get("field"))
    if expression_type == "ref":
        block_name = clean_cell(expression.get("name"))
        if detection_blocks and block_name in detection_blocks:
            return score_expression_to_text(detection_blocks[block_name]["expression"], detection_blocks)
        return block_name
    if expression_type == "missing_ref":
        return f"<missing {clean_cell(expression.get('name'))}>"
    if expression_type == "not":
        return f"NOT {score_expression_to_text(expression.get('child', {'type': 'empty'}), detection_blocks)}"
    if expression_type in {"and", "or"}:
        operator = f" {expression_type.upper()} "
        return "(" + operator.join(score_expression_to_text(child, detection_blocks) for child in expression.get("children", [])) + ")"
    return "<empty>"


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    output: list[str] = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        output.append(value)
    return output


def build_data_component_candidates(
    rule: dict[str, Any],
    field_score_rows: list[dict[str, Any]],
    event_context_rows: list[dict[str, Any]],
) -> list[str]:
    candidates: list[str] = []

    for row in event_context_rows:
        append_unique_component(candidates, row.get("ATT&CK Data Component (v19)"))
    for row in field_score_rows:
        append_unique_component(candidates, row.get("Data Component"))

    # Fallback for Sigma categories whose names already mirror ATT&CK data components.
    append_unique_component(
        candidates,
        sigma_category_to_component_label(rule["rule_logsource_category"] or rule["rule_logsource_service"]),
    )
    return candidates


def append_unique_component(target: list[str], value: Any) -> None:
    component = clean_cell(value)
    if not component or normalize_data_component(component) == "no equivalent component":
        return
    if normalize_data_component(component) not in {normalize_data_component(existing) for existing in target}:
        target.append(component)


def sigma_category_to_component_label(value: Any) -> str:
    text = normalize_phrase(value)
    return text.title() if text else ""


def build_implementation_rows(
    rule: dict[str, Any],
    attack_rows: list[dict[str, Any]],
    data_component_candidates: list[str],
) -> list[dict[str, Any]]:
    technique_ids = set(rule["attack_technique_ids"])
    component_tokens = {normalize_data_component(component) for component in data_component_candidates}
    if not component_tokens or not technique_ids:
        return []

    output_rows: list[dict[str, Any]] = []
    for row in attack_rows:
        technique_id = clean_cell(row.get("ATT&CK Technique ID"))
        if technique_id not in technique_ids:
            continue
        row_component_tokens = {
            normalize_data_component(component)
            for component in split_multi_value(row.get("ATT&CK Data Component (v19)"), separators=(";", ","))
        }
        if not component_tokens.intersection(row_component_tokens):
            continue
        output_rows.append(
            {
                "ATT&CK Technique ID": technique_id,
                "ATT&CK Technique Name": clean_cell(row.get("ATT&CK Technique Name")),
                "Technique Implementation name": clean_cell(row.get("Technique Implementation")),
                "Implementation Step Number": clean_cell(row.get("Implementation Step Number")),
                "Implementation Step": clean_cell(row.get("Implementation Step")),
                "System Interaction": clean_cell(row.get("System Interaction")),
                "ATT&CK Data Component (v19)": clean_cell(row.get("ATT&CK Data Component (v19)")),
                "Detection Observables": clean_cell(row.get("Detection Observable")),
                "Analytic Files": rule["analytic_file"],
                "Analytic Titles": rule["rule_title"],
            }
        )
    return output_rows


def consolidate_implementation_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, ...], dict[str, Any]] = {}
    analytic_files_by_key: dict[tuple[str, ...], list[str]] = {}
    analytic_titles_by_key: dict[tuple[str, ...], list[str]] = {}

    key_headers = [
        "ATT&CK Technique ID",
        "ATT&CK Technique Name",
        "Technique Implementation name",
        "Implementation Step Number",
        "Implementation Step",
        "System Interaction",
        "ATT&CK Data Component (v19)",
        "Detection Observables",
    ]
    for row in rows:
        key = tuple(clean_cell(row.get(header)) for header in key_headers)
        if key not in grouped:
            grouped[key] = {header: clean_cell(row.get(header)) for header in IMPLEMENTATION_HEADERS}
            analytic_files_by_key[key] = []
            analytic_titles_by_key[key] = []
        append_unique(analytic_files_by_key[key], clean_cell(row.get("Analytic Files")))
        append_unique(analytic_titles_by_key[key], clean_cell(row.get("Analytic Titles")))

    output_rows = []
    for key in sorted(grouped):
        row = grouped[key]
        row["Analytic Files"] = "; ".join(analytic_files_by_key[key])
        row["Analytic Titles"] = "; ".join(analytic_titles_by_key[key])
        output_rows.append(row)
    return output_rows


def build_implementation_catalog_rows(
    analytic_summaries: list[dict[str, Any]],
    attack_rows: list[dict[str, Any]],
    matched_implementation_rows: list[dict[str, Any]],
    technique_coverage_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    technique_ids = {
        technique_id
        for summary in analytic_summaries
        for technique_id in summary["rule"]["attack_technique_ids"]
    }
    coverage_by_technique = {
        clean_cell(row.get("ATT&CK Technique ID")): row for row in technique_coverage_rows
    }
    matched_by_technique_implementation_and_step: dict[tuple[str, str, str], dict[str, list[str]]] = {}
    for row in matched_implementation_rows:
        key = (
            clean_cell(row.get("ATT&CK Technique ID")),
            normalize_token(row.get("Technique Implementation name")),
            normalize_token(row.get("Implementation Step")),
        )
        entry = matched_by_technique_implementation_and_step.setdefault(key, {"files": [], "titles": []})
        for analytic_file in split_multi_value(row.get("Analytic Files"), separators=(";",)):
            append_unique(entry["files"], analytic_file)
        for analytic_title in split_multi_value(row.get("Analytic Titles"), separators=(";",)):
            append_unique(entry["titles"], analytic_title)

    output_rows: list[dict[str, Any]] = []
    seen: set[tuple[str, ...]] = set()
    for row in attack_rows:
        technique_id = clean_cell(row.get("ATT&CK Technique ID"))
        if technique_id not in technique_ids:
            continue

        implementation_name = clean_cell(row.get("Technique Implementation"))
        implementation_step = clean_cell(row.get("Implementation Step"))
        match_key = (technique_id, normalize_token(implementation_name), normalize_token(implementation_step))
        match_info = matched_by_technique_implementation_and_step.get(match_key, {"files": [], "titles": []})
        covered = "yes" if match_info["files"] or match_info["titles"] else "no"
        coverage = coverage_by_technique.get(technique_id, {})
        output_row = {
            "ATT&CK Technique ID": technique_id,
            "ATT&CK Technique Name": clean_cell(row.get("ATT&CK Technique Name")),
            "Technique Implementation name": implementation_name,
            "Implementation Step Number": clean_cell(row.get("Implementation Step Number")),
            "Implementation Step": implementation_step,
            "System Interaction": clean_cell(row.get("System Interaction")),
            "ATT&CK Data Component (v19)": clean_cell(row.get("ATT&CK Data Component (v19)")),
            "Detection Observables": clean_cell(row.get("Detection Observable")),
            "Covered By Analytics": covered,
            "Matched Implementation Count": clean_cell(coverage.get("Matched Implementation Count")),
            "Total Implementation Count": clean_cell(coverage.get("Total Implementation Count")),
            "Implementation Coverage Ratio": clean_cell(coverage.get("Implementation Coverage Ratio")),
            "Implementation Coverage Percent": clean_cell(coverage.get("Implementation Coverage Percent")),
            "Analytic Files": "; ".join(match_info["files"]),
            "Analytic Titles": "; ".join(match_info["titles"]),
        }
        row_key = tuple(clean_cell(output_row.get(header)) for header in IMPLEMENTATION_HEADERS)
        if row_key in seen:
            continue
        seen.add(row_key)
        output_rows.append(output_row)
    return output_rows


def build_technique_coverage_rows(
    analytic_summaries: list[dict[str, Any]],
    attack_rows: list[dict[str, Any]],
    implementation_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    technique_ids = sorted(
        {
            technique_id
            for summary in analytic_summaries
            for technique_id in summary["rule"]["attack_technique_ids"]
        }
    )
    tagged_analytic_files_by_technique: dict[str, list[str]] = {technique_id: [] for technique_id in technique_ids}
    tagged_analytic_titles_by_technique: dict[str, list[str]] = {technique_id: [] for technique_id in technique_ids}
    for summary in analytic_summaries:
        rule = summary["rule"]
        for technique_id in rule["attack_technique_ids"]:
            append_unique(tagged_analytic_files_by_technique.setdefault(technique_id, []), rule["analytic_file"])
            append_unique(tagged_analytic_titles_by_technique.setdefault(technique_id, []), rule["rule_title"])

    all_implementations_by_technique: dict[str, list[str]] = {technique_id: [] for technique_id in technique_ids}
    technique_names: dict[str, str] = {}
    for row in attack_rows:
        technique_id = clean_cell(row.get("ATT&CK Technique ID"))
        if technique_id not in all_implementations_by_technique:
            continue
        if technique_id not in technique_names and clean_cell(row.get("ATT&CK Technique Name")):
            technique_names[technique_id] = clean_cell(row.get("ATT&CK Technique Name"))
        append_unique(all_implementations_by_technique[technique_id], clean_cell(row.get("Technique Implementation")))

    matched_implementations_by_technique: dict[str, list[str]] = {technique_id: [] for technique_id in technique_ids}
    matched_analytic_files_by_technique: dict[str, list[str]] = {technique_id: [] for technique_id in technique_ids}
    matched_analytic_titles_by_technique: dict[str, list[str]] = {technique_id: [] for technique_id in technique_ids}
    for row in implementation_rows:
        technique_id = clean_cell(row.get("ATT&CK Technique ID"))
        if technique_id not in matched_implementations_by_technique:
            continue
        append_unique(matched_implementations_by_technique[technique_id], clean_cell(row.get("Technique Implementation name")))
        for analytic_file in split_multi_value(row.get("Analytic Files"), separators=(";",)):
            append_unique(matched_analytic_files_by_technique[technique_id], analytic_file)
        for analytic_title in split_multi_value(row.get("Analytic Titles"), separators=(";",)):
            append_unique(matched_analytic_titles_by_technique[technique_id], analytic_title)
        if technique_id not in technique_names and clean_cell(row.get("ATT&CK Technique Name")):
            technique_names[technique_id] = clean_cell(row.get("ATT&CK Technique Name"))

    coverage_rows: list[dict[str, Any]] = []
    for technique_id in technique_ids:
        total_count = len(all_implementations_by_technique.get(technique_id, []))
        matched_count = len(matched_implementations_by_technique.get(technique_id, []))
        coverage_rows.append(
            {
                "ATT&CK Technique ID": technique_id,
                "ATT&CK Technique Name": technique_names.get(technique_id, ""),
                "Matched Implementation Count": matched_count,
                "Total Implementation Count": total_count,
                "Implementation Coverage Ratio": f"{matched_count}/{total_count}",
                "Implementation Coverage Percent": coverage_percent(matched_count, total_count),
                "Matched Technique Implementations": "; ".join(matched_implementations_by_technique.get(technique_id, [])),
                "All Technique Implementations": "; ".join(all_implementations_by_technique.get(technique_id, [])),
                "Analytic Files": "; ".join(matched_analytic_files_by_technique.get(technique_id, [])),
                "Analytic Titles": "; ".join(matched_analytic_titles_by_technique.get(technique_id, [])),
                "Tagged Analytic Files": "; ".join(tagged_analytic_files_by_technique.get(technique_id, [])),
                "Tagged Analytic Titles": "; ".join(tagged_analytic_titles_by_technique.get(technique_id, [])),
            }
        )
    return coverage_rows


def add_coverage_to_implementation_rows(
    implementation_rows: list[dict[str, Any]],
    technique_coverage_rows: list[dict[str, Any]],
) -> None:
    coverage_by_technique = {
        clean_cell(row.get("ATT&CK Technique ID")): row for row in technique_coverage_rows
    }
    for row in implementation_rows:
        coverage = coverage_by_technique.get(clean_cell(row.get("ATT&CK Technique ID")), {})
        row["Matched Implementation Count"] = clean_cell(coverage.get("Matched Implementation Count"))
        row["Total Implementation Count"] = clean_cell(coverage.get("Total Implementation Count"))
        row["Implementation Coverage Ratio"] = clean_cell(coverage.get("Implementation Coverage Ratio"))
        row["Implementation Coverage Percent"] = clean_cell(coverage.get("Implementation Coverage Percent"))


def coverage_percent(matched_count: int, total_count: int) -> str:
    if total_count == 0:
        return "0.0%"
    return f"{(matched_count / total_count) * 100:.1f}%"


def scoring_context_matches(rule: dict[str, Any], scoring_row: dict[str, Any]) -> bool:
    return scoring_logsource_matches(rule, scoring_row) and scoring_event_id_matches(rule, scoring_row)


def scoring_logsource_matches(rule: dict[str, Any], scoring_row: dict[str, Any]) -> bool:
    checks = (
        ("rule_logsource_category", SCORING_CATEGORY_COLUMNS),
        ("rule_logsource_service", SCORING_SERVICE_COLUMNS),
        ("rule_logsource_product", SCORING_PRODUCT_COLUMNS),
    )
    for rule_key, column_aliases in checks:
        rule_value = clean_cell(rule.get(rule_key))
        if not rule_value:
            continue
        if not row_has_any_column(scoring_row, column_aliases):
            continue
        if not multi_value_cell_matches(rule_value, scoring_row_value(scoring_row, column_aliases)):
            return False
    return True


def scoring_event_id_matches(rule: dict[str, Any], scoring_row: dict[str, Any]) -> bool:
    event_ids = set(rule_event_ids(rule))
    if not event_ids:
        return True
    if not row_has_any_column(scoring_row, SCORING_EVENT_ID_COLUMNS):
        return False

    workbook_event_ids = {
        normalize_event_id(token)
        for token in split_multi_value(scoring_row_value(scoring_row, SCORING_EVENT_ID_COLUMNS), separators=(";", ","))
    }
    workbook_event_ids.discard("")
    return bool(event_ids.intersection(workbook_event_ids))


def rule_event_ids(rule: dict[str, Any]) -> list[str]:
    event_ids: list[str] = []
    for field, values in rule.get("detection_values", {}).items():
        if not is_event_id_field(field):
            continue
        for value in values:
            for token in split_multi_value(value, separators=(";", ",")):
                append_unique(event_ids, normalize_event_id(token))
    return event_ids


def is_event_id_field(value: Any) -> bool:
    return normalize_field_name(value) == "eventid"


def normalize_event_id(value: Any) -> str:
    text = clean_cell(value)
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def row_has_any_column(row: dict[str, Any], column_aliases: tuple[str, ...]) -> bool:
    return matching_row_column(row, column_aliases) != ""


def scoring_row_value(row: dict[str, Any], column_aliases: tuple[str, ...]) -> str:
    column = matching_row_column(row, column_aliases)
    if column:
        return clean_cell(row.get(column))
    return ""


def matching_row_column(row: dict[str, Any], column_aliases: tuple[str, ...]) -> str:
    for column in column_aliases:
        if column in row:
            return column

    normalized_aliases = {normalize_header(alias) for alias in column_aliases}
    for column in row:
        if normalize_header(column) in normalized_aliases:
            return column
    return ""


def multi_value_cell_matches(rule_value: Any, workbook_value: Any) -> bool:
    rule_token = normalize_token(rule_value)
    if not rule_token:
        return True
    workbook_tokens = split_multi_value(workbook_value, separators=(";", ","))
    return rule_token in {normalize_token(token) for token in workbook_tokens}


def logsource_matches(rule: dict[str, Any], workbook_logsource: Any) -> bool:
    workbook_token = normalize_token(workbook_logsource)
    if not workbook_token:
        return True

    rule_tokens = {
        normalize_token(rule["rule_logsource_category"]),
        normalize_token(rule["rule_logsource_service"]),
        normalize_token(rule["rule_logsource_product"]),
    }
    rule_tokens.discard("")
    if workbook_token in rule_tokens:
        return True

    return workbook_token == "windows event log" and "windows" in rule_tokens


def field_matches_any(workbook_field: Any, detection_fields: set[str]) -> bool:
    workbook_token = normalize_field_name(workbook_field)
    if not workbook_token:
        return False
    return any(workbook_token == normalize_field_name(field) for field in detection_fields)


def same_text(left: Any, right: Any) -> bool:
    return normalize_token(left) == normalize_token(right)


def split_multi_value(value: Any, separators: tuple[str, ...]) -> list[str]:
    text = clean_cell(value)
    if not text:
        return []
    pattern = "|".join(re.escape(separator) for separator in separators)
    return [part.strip() for part in re.split(pattern, text) if part and part.strip()]


def normalize_token(value: Any) -> str:
    return clean_cell(value).strip().lower()


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_token(value))


def normalize_phrase(value: Any) -> str:
    text = clean_cell(value).strip().lower()
    text = re.sub(r"[_-]+", " ", text)
    return re.sub(r"\s+", " ", text)


def normalize_data_component(value: Any) -> str:
    return normalize_phrase(value)


def normalize_field_name(value: Any) -> str:
    text = normalize_token(value)
    for prefix in ("system.", "eventdata."):
        if text.startswith(prefix):
            return text.removeprefix(prefix)
    return text


def clean_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def build_summary_rows(
    analytic_summaries: list[dict[str, Any]],
    scoring_sheet: str,
    attack_sheet: str,
    event_source: str,
    event_context_rows: list[dict[str, Any]],
    field_rows: list[dict[str, Any]],
    implementation_rows: list[dict[str, Any]],
    technique_coverage_rows: list[dict[str, Any]],
    analytic_score_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    technique_ids = sorted(
        {
            technique_id
            for summary in analytic_summaries
            for technique_id in summary["rule"]["attack_technique_ids"]
        }
    )
    rows = [
        {"Metric": "Analytics Processed", "Value": len(analytic_summaries)},
        {"Metric": "Analytic Files", "Value": "; ".join(summary["rule"]["analytic_file"] for summary in analytic_summaries)},
        {"Metric": "ATT&CK Technique IDs", "Value": ", ".join(technique_ids)},
        {"Metric": "Scoring Sheet", "Value": scoring_sheet},
        {"Metric": "ATT&CK Sheet", "Value": attack_sheet},
        {"Metric": "Event/Field Source", "Value": event_source},
        {"Metric": "Event Context Match Count", "Value": len(event_context_rows)},
        {"Metric": "Implementation Catalog Row Count", "Value": len(implementation_rows)},
        {"Metric": "Covered Implementation Row Count", "Value": sum(1 for row in implementation_rows if row.get("Covered By Analytics") == "yes")},
        {"Metric": "Technique Coverage Row Count", "Value": len(technique_coverage_rows)},
        {"Metric": "Analytic Score Row Count", "Value": len(analytic_score_rows)},
        {"Metric": "Field Score Row Count", "Value": len(field_rows)},
        {"Metric": "Field Score Match Count", "Value": sum(1 for row in field_rows if row["Matched"] == "yes")},
    ]
    for analytic_score in analytic_score_rows:
        rows.append(
            {
                "Metric": f"Score: {analytic_score['Analytic File']}",
                "Value": format_summary_score_value(analytic_score),
            }
        )
    for coverage in technique_coverage_rows:
        rows.append(
            {
                "Metric": f"Coverage: {coverage['ATT&CK Technique ID']}",
                "Value": (
                    f"{coverage['Matched Implementation Count']}/"
                    f"{coverage['Total Implementation Count']} implementations "
                    f"({coverage['Implementation Coverage Percent']})"
                ),
            }
        )
    for summary in analytic_summaries:
        rule = summary["rule"]
        rows.append(
            {
                "Metric": f"Analytic: {rule['analytic_file']}",
                "Value": (
                    f"{summary['field_score_match_count']}/{summary['field_score_row_count']} field scores; "
                    f"{summary['implementation_match_count']} implementation matches; "
                    f"components: {', '.join(summary['attack_data_component_candidates']) or '(none)'}"
                ),
            }
        )
    return rows


def format_summary_score_value(analytic_score: dict[str, Any]) -> str:
    robustness = clean_cell(analytic_score.get("Robustness Score"))
    precision = clean_cell(analytic_score.get("Precision Score"))
    if robustness and precision:
        return f"Robustness: {robustness}, Precision: {precision}"
    missing = clean_cell(analytic_score.get("Missing Scores"))
    return f"unscored ({missing})" if missing else "unscored"


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(
    path: Path,
    implementation_rows: list[dict[str, Any]],
    field_score_rows: list[dict[str, Any]],
    event_context_rows: list[dict[str, Any]],
    technique_coverage_rows: list[dict[str, Any]],
    analytic_score_rows: list[dict[str, Any]],
    summary_rows: list[dict[str, Any]],
) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    write_sheet(summary_sheet, ["Metric", "Value"], summary_rows)
    write_sheet(workbook.create_sheet("Implementations"), IMPLEMENTATION_HEADERS, implementation_rows)
    write_sheet(workbook.create_sheet("Technique Coverage"), TECHNIQUE_COVERAGE_HEADERS, technique_coverage_rows)
    write_sheet(workbook.create_sheet("Analytic Scores"), ANALYTIC_SCORE_HEADERS, analytic_score_rows)
    write_sheet(workbook.create_sheet("Field Scores"), FIELD_SCORE_HEADERS, field_score_rows)
    write_sheet(workbook.create_sheet("OCSF Context"), OCSF_CONTEXT_HEADERS, event_context_rows)
    workbook.save(path)


def write_sheet(worksheet: Any, headers: list[str], rows: list[dict[str, Any]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def autosize_columns(worksheet: Any) -> None:
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            max_length = max(max_length, len(clean_cell(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def write_summary_json(
    path: Path,
    analytic_summaries: list[dict[str, Any]],
    scoring_sheet: str,
    attack_sheet: str,
    event_source: str,
    event_context_rows: list[dict[str, Any]],
    field_score_rows: list[dict[str, Any]],
    implementation_rows: list[dict[str, Any]],
    technique_coverage_rows: list[dict[str, Any]],
    analytic_score_rows: list[dict[str, Any]],
) -> None:
    technique_ids = sorted(
        {
            technique_id
            for summary in analytic_summaries
            for technique_id in summary["rule"]["attack_technique_ids"]
        }
    )
    payload = {
        "analytics_processed": len(analytic_summaries),
        "technique_ids": technique_ids,
        "scoring_sheet": scoring_sheet,
        "attack_sheet": attack_sheet,
        "event_source": event_source,
        "event_context_match_count": len(event_context_rows),
        "implementation_match_count": len(implementation_rows),
        "technique_coverage": technique_coverage_rows,
        "analytic_scores": analytic_score_rows,
        "field_score_row_count": len(field_score_rows),
        "field_score_match_count": sum(1 for row in field_score_rows if row["Matched"] == "yes"),
        "analytics": analytic_summaries,
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
